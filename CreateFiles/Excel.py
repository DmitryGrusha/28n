import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

def create_and_fill_excel(name: str, sex: str, birthday: str, age: str, doctors: [str], cabinets: [str]):

    data = {
        'Категория': ['ФИО', 'Пол', 'Дата рождения', 'Возраст', 'Специалисты', 'Кабинеты'],
        'Значение': [name, sex, birthday, age, ', '.join(doctors), ', '.join(cabinets)]
    }

    df = pd.DataFrame(data)


    filename = f'{name}.xlsx'
    df.to_excel(filename, index=False, header=False, engine='openpyxl')


    wb = load_workbook(filename)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    bold_font = Font(bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Выравнивание по левому краю
            if cell.row == 1:  # Применение жирного шрифта только к первой строке
                cell.font = bold_font

    wb.save(filename)
    print("File successfully saved with vertical alignment and adjusted column widths")